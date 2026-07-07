"""
config/item_mapping.csv 에 '수출입_hs_code_종목_매핑_세부 정리본.xlsx'의 HS코드/관련종목 데이터를
안전하게 병합하는 1회성(재실행 가능) 스크립트.

원칙:
- item_mapping.csv의 기존 값이 있는 셀은 절대 덮어쓰지 않는다. 비어있는
  related_companies/hs_code 칸만 채운다.
- trade_history_long.csv에는 있지만 item_mapping.csv에 아직 없는 품목은
  utils_data.load_item_mapping()과 동일한 규칙(category=품목명 첫 '_' 앞 토큰,
  related_companies/hs_code는 빈 값)으로 새 행을 추가한 뒤, 위 규칙대로 xlsx 데이터를 채운다.
- config/item_hscode_detail.csv(세부품목명, hs_code long-format)는 매번 xlsx 내용으로
  통째로 재생성한다 (수기 편집 대상이 아니라 xlsx의 파생 데이터이므로).
- xlsx의 S열(카테고리 분류)은 D열과 행 단위로 대응되지 않는 깨진 데이터라 사용하지 않는다.

수동 확정 사항 (2026-07-07 사용자 확인):
- '반도체_테스트소켓'(trade_history)과 '반도체_테스트 핀/포고핀'(xlsx)은 실제로 다른
  부품이라 별개 품목으로 유지한다 (자동 연결하지 않음 -> 코드에서 별도 처리 불필요,
  xlsx 쪽에 애초에 '반도체_테스트소켓'이라는 이름이 없으므로 자연히 매핑 안 됨).
- xlsx의 '전기전자_디스플레이_OLED'(세부품목 10개: 노트북/스마트폰/TV/모니터/기타 x2세트)는
  trade_history의 '전기전자_디스플레이_OLED 전체' 품목에 매핑한다. '..OLED 모바일'/
  '..OLED_TV'는 매핑하지 않고 빈 값으로 남긴다.

사용법:
    python import_item_mapping.py            # dry-run: 변경 예정 내역만 출력, 파일 쓰지 않음
    python import_item_mapping.py --apply     # 실제로 item_mapping.csv / item_hscode_detail.csv 갱신
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).parent
XLSX_PATH = Path(r"C:\Users\김현준\Desktop\수출입_hs_code_종목_매핑_세부 정리본.xlsx")
SHEET_NAME = "카테고리분류(원본)"
HISTORY_PATH = BASE_DIR / "trade_history_long.csv"
MAPPING_PATH = BASE_DIR / "config" / "item_mapping.csv"
DETAIL_PATH = BASE_DIR / "config" / "item_hscode_detail.csv"
MAPPING_COLUMNS = ["item_name", "category", "related_companies", "hs_code"]

# xlsx 품목명(키) -> trade_history_long.csv 품목명(값)으로 강제 치환. 사용자 확정 사항만 담는다.
ITEM_NAME_OVERRIDES = {
    "전기전자_디스플레이_OLED": "전기전자_디스플레이_OLED 전체",
}

DELIM = "; "


def load_xlsx_items() -> dict[str, dict]:
    """xlsx를 파싱해 {품목명: {"companies": [...], "hs_details": [(detail_name, hs_code), ...]}} 로 정리."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    for r in range(4, ws.max_row + 1):
        item = ws.cell(row=r, column=4).value
        if item is None:
            continue
        item = str(item).strip()
        detail = ws.cell(row=r, column=5).value
        detail = detail.strip() if isinstance(detail, str) else ""
        hs = ws.cell(row=r, column=6).value
        hs_str = str(int(hs)) if isinstance(hs, float) else str(hs).strip()
        companies = []
        for c in range(7, 16):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                companies.append(str(v).strip())
        rows.append((item, detail, hs_str, companies))

    by_item: dict[str, dict] = defaultdict(lambda: {"companies": [], "hs_details": []})
    for item, detail, hs_str, companies in rows:
        item = ITEM_NAME_OVERRIDES.get(item, item)
        entry = by_item[item]
        if not entry["companies"]:
            entry["companies"] = companies
        seen_hs = {h for _, h in entry["hs_details"]}
        if hs_str not in seen_hs:
            entry["hs_details"].append((detail, hs_str))
    return dict(by_item)


def build_detail_df(xlsx_items: dict[str, dict]) -> pd.DataFrame:
    rows = []
    for item, data in xlsx_items.items():
        for detail, hs in data["hs_details"]:
            rows.append({"item_name": item, "detail_name": detail, "hs_code": hs})
    return pd.DataFrame(rows, columns=["item_name", "detail_name", "hs_code"]).sort_values(
        ["item_name", "hs_code"]
    ).reset_index(drop=True)


def load_or_init_mapping(history_items: list[str]) -> tuple[pd.DataFrame, list[str]]:
    """기존 item_mapping.csv를 읽고, trade_history에만 있는 품목은 새 행으로 추가.
    반환: (mapping df, 새로 추가된 품목명 리스트)"""
    mapping = pd.read_csv(MAPPING_PATH, dtype=str).fillna("")
    for col in MAPPING_COLUMNS:
        if col not in mapping.columns:
            mapping[col] = ""

    existing = set(mapping["item_name"])
    new_items = [i for i in history_items if i not in existing]
    if new_items:
        new_rows = pd.DataFrame(
            {
                "item_name": new_items,
                "category": [i.split("_")[0] for i in new_items],
                "related_companies": ["" for _ in new_items],
                "hs_code": ["" for _ in new_items],
            }
        )
        mapping = pd.concat([mapping, new_rows], ignore_index=True)
    return mapping[MAPPING_COLUMNS], new_items


def main():
    apply = "--apply" in sys.argv

    history_items = sorted(pd.read_csv(HISTORY_PATH)["품목명"].dropna().unique().tolist())
    xlsx_items = load_xlsx_items()
    mapping, new_items = load_or_init_mapping(history_items)

    fills = []  # (item_name, field, old, new)
    for idx, row in mapping.iterrows():
        item = row["item_name"]
        data = xlsx_items.get(item)
        if data is None:
            continue
        if row["related_companies"] == "" and data["companies"]:
            new_val = DELIM.join(data["companies"])
            fills.append((item, "related_companies", "", new_val))
            mapping.at[idx, "related_companies"] = new_val
        if row["hs_code"] == "" and data["hs_details"]:
            new_val = DELIM.join(h for _, h in data["hs_details"])
            fills.append((item, "hs_code", "", new_val))
            mapping.at[idx, "hs_code"] = new_val

    detail_df = build_detail_df(xlsx_items)

    print(f"xlsx 품목 수: {len(xlsx_items)}")
    print(f"trade_history 품목 수: {len(history_items)}")
    print(f"item_mapping.csv에 새로 추가될 품목: {len(new_items)}개")
    for i in new_items:
        print(f"  + {i}")
    print()
    print(f"채워질 필드: {len(fills)}건 (품목 {len(set(f[0] for f in fills))}개)")
    for item, field, old, new in fills:
        shown = new if len(new) <= 80 else new[:77] + "..."
        print(f"  [{item}] {field}: '' -> '{shown}'")
    print()
    print(f"item_hscode_detail.csv 생성 예정 행 수: {len(detail_df)} (품목 {detail_df['item_name'].nunique()}개)")

    unmapped_history = [i for i in history_items if i not in xlsx_items and mapping.set_index("item_name").loc[i, "hs_code"] == "" ]
    print()
    print(f"xlsx 매핑 데이터가 없어 빈 값으로 남는 trade_history 품목: {len(unmapped_history)}개")
    for i in unmapped_history:
        print(f"  - {i}")

    if not apply:
        print()
        print("[DRY-RUN] 실제 파일은 변경하지 않았습니다. 반영하려면 --apply 옵션으로 재실행하세요.")
        return

    mapping.to_csv(MAPPING_PATH, index=False)
    detail_df.to_csv(DETAIL_PATH, index=False)
    print()
    print(f"[APPLY] {MAPPING_PATH} 갱신 완료 ({len(fills)}건 채움, 신규 {len(new_items)}행 추가)")
    print(f"[APPLY] {DETAIL_PATH} 생성/갱신 완료 ({len(detail_df)}행)")


if __name__ == "__main__":
    main()
